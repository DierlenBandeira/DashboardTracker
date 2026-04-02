import io
import json
import mimetypes
import os
import re
import time
import threading
import uuid
from bisect import bisect_left
from collections import deque
from datetime import datetime, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, unquote, parse_qs
from pathlib import Path

import requests
from openpyxl import load_workbook, Workbook

# -----------------------------
# Config
# -----------------------------
WIALON_HOST = os.getenv("WIALON_HOST", "https://hst-api.wialon.us").rstrip("/")
POLL_SECONDS = float(os.getenv("POLL_SECONDS", "10"))

HTTP_HOST = os.getenv("HTTP_HOST", "0.0.0.0")
HTTP_PORT = int(os.getenv("PORT", os.getenv("HTTP_PORT", "8790")))
TIMEOUT_SECONDS = float(os.getenv("TIMEOUT_SECONDS", "12"))
STALE_THRESHOLD_SEC = float(os.getenv("STALE_THRESHOLD_SEC", "180"))
LIVE_BUFFER_SIZE = int(os.getenv("LIVE_BUFFER_SIZE", "720"))

BASE_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = BASE_DIR.parent / "frontend"
TEMP_TREATMENT_DIR = BASE_DIR / "_treatment_jobs"
TEMP_TREATMENT_DIR.mkdir(parents=True, exist_ok=True)
TEMP_TREATMENT_INBOX_DIR = TEMP_TREATMENT_DIR / "inbox"
TEMP_TREATMENT_INBOX_DIR.mkdir(parents=True, exist_ok=True)

TREATMENT_JOBS_LOCK = threading.Lock()
TREATMENT_JOBS = {}
TREATMENT_STATUS_LOCK = threading.Lock()
TREATMENT_STATUS = {}
PREVIEW_LIMIT = 200

HTML_PATH = FRONTEND_DIR / "index.html"
CSS_PATH = FRONTEND_DIR / "dashboard.css"
LOGO_PATH = FRONTEND_DIR / "logo.png"

CORS_ALLOWED_ORIGINS = [
    origin.strip()
    for origin in os.getenv(
        "CORS_ALLOWED_ORIGINS",
        "http://localhost:8790,http://127.0.0.1:8790,https://dashboard-tracker-weld.vercel.app",
    ).split(",")
    if origin.strip()
]

CORS_ALLOW_METHODS = "GET, POST, OPTIONS"
CORS_ALLOW_HEADERS = "Content-Type, X-Filename, X-Treatment-Config"

SENSOR_IDX = {
    "velocidade": 8,
    "rpm": 9,
    "consumido": 6,
    "temperatura_motor": 13,
    "pct_acelerado": 31,
    "ar_cond": 38,
    "altitude": 46,
}

AVL_LOCK = threading.Lock()
AVL_CURSOR_TM = int(time.time()) - 120

SID_LOCK = threading.Lock()
SID = os.getenv("WIALON_SID", "").strip()

UNIT_LOCK = threading.Lock()
ITEM_ID = int(os.getenv("WIALON_ITEM_ID", "401833488"))

STATE_LOCK = threading.Lock()
STATE = {
    "ok": False,
    "last_update_epoch": None,
    "error": None,
    "data": None,
    "raw": None,
    "ev_raw": None,
}

LIVE_BUFFER_LOCK = threading.Lock()
LIVE_BUFFER_BY_UNIT = {}

LAST_MSG_TM_LOCK = threading.Lock()
LAST_MSG_TM_BY_UNIT = {}

LAST_SIGNATURE_LOCK = threading.Lock()
LAST_SIGNATURE_BY_UNIT = {}

LAST_PLOT_KIND_LOCK = threading.Lock()
LAST_PLOT_KIND_BY_UNIT = {}

BOOTSTRAP_DONE_LOCK = threading.Lock()
BOOTSTRAP_DONE_BY_UNIT = {}


# -----------------------------
# Utils
# -----------------------------
def normalize_resp(resp):
    if isinstance(resp, list):
        if not resp:
            return {}
        return resp[0] if isinstance(resp[0], dict) else {}
    return resp if isinstance(resp, dict) else {}


def to_num(v):
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return None


def first_not_none(*values):
    for v in values:
        if v is not None:
            return v
    return None


def normalize_bool_like(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return 1 if v else 0
    n = to_num(v)
    if n is not None:
        return 1 if n != 0 else 0
    s = str(v).strip().lower()
    if s in {"ligado", "on", "true", "sim"}:
        return 1
    if s in {"desligado", "off", "false", "nao", "não"}:
        return 0
    return None


def sensor_field(raw_value, display_value=None):
    return {"raw": raw_value, "display": display_value}


# -----------------------------
# State helpers
# -----------------------------
def get_sid():
    with SID_LOCK:
        return SID


def set_sid(new_sid: str):
    global SID
    with SID_LOCK:
        SID = (new_sid or "").strip()


def get_item_id():
    with UNIT_LOCK:
        return ITEM_ID


def set_item_id(new_id: int):
    global ITEM_ID
    with UNIT_LOCK:
        ITEM_ID = int(new_id)


def get_live_buffer(item_id: int):
    with LIVE_BUFFER_LOCK:
        if item_id not in LIVE_BUFFER_BY_UNIT:
            LIVE_BUFFER_BY_UNIT[item_id] = deque(maxlen=LIVE_BUFFER_SIZE)
        return LIVE_BUFFER_BY_UNIT[item_id]


def append_live_sample(item_id: int, sample: dict):
    buf = get_live_buffer(item_id)
    with LIVE_BUFFER_LOCK:
        buf.append(sample)


def list_live_samples(item_id: int, limit: int | None = None):
    buf = get_live_buffer(item_id)
    with LIVE_BUFFER_LOCK:
        data = list(buf)
    if limit is not None and limit > 0:
        return data[-limit:]
    return data


def clear_live_buffer(item_id: int):
    with LIVE_BUFFER_LOCK:
        LIVE_BUFFER_BY_UNIT.pop(item_id, None)


def get_last_msg_tm(item_id: int):
    with LAST_MSG_TM_LOCK:
        return LAST_MSG_TM_BY_UNIT.get(item_id)


def set_last_msg_tm(item_id: int, value):
    with LAST_MSG_TM_LOCK:
        if value is None:
            LAST_MSG_TM_BY_UNIT.pop(item_id, None)
        else:
            LAST_MSG_TM_BY_UNIT[item_id] = int(value)


def get_last_signature(item_id: int):
    with LAST_SIGNATURE_LOCK:
        return LAST_SIGNATURE_BY_UNIT.get(item_id)


def set_last_signature(item_id: int, value):
    with LAST_SIGNATURE_LOCK:
        if value is None:
            LAST_SIGNATURE_BY_UNIT.pop(item_id, None)
        else:
            LAST_SIGNATURE_BY_UNIT[item_id] = value


def get_last_plot_kind(item_id: int):
    with LAST_PLOT_KIND_LOCK:
        return LAST_PLOT_KIND_BY_UNIT.get(item_id)


def set_last_plot_kind(item_id: int, value):
    with LAST_PLOT_KIND_LOCK:
        if value is None:
            LAST_PLOT_KIND_BY_UNIT.pop(item_id, None)
        else:
            LAST_PLOT_KIND_BY_UNIT[item_id] = value


def get_bootstrap_done(item_id: int):
    with BOOTSTRAP_DONE_LOCK:
        return bool(BOOTSTRAP_DONE_BY_UNIT.get(item_id))


def set_bootstrap_done(item_id: int, done: bool):
    with BOOTSTRAP_DONE_LOCK:
        if done:
            BOOTSTRAP_DONE_BY_UNIT[item_id] = True
        else:
            BOOTSTRAP_DONE_BY_UNIT.pop(item_id, None)


def reset_runtime_state(item_id: int):
    global AVL_CURSOR_TM

    with AVL_LOCK:
        AVL_CURSOR_TM = int(time.time()) - 120

    clear_live_buffer(item_id)
    set_last_msg_tm(item_id, None)
    set_last_signature(item_id, None)
    set_last_plot_kind(item_id, None)
    set_bootstrap_done(item_id, False)

    with STATE_LOCK:
        STATE["ok"] = False
        STATE["error"] = None
        STATE["data"] = None
        STATE["raw"] = None
        STATE["ev_raw"] = None


# -----------------------------
# Static serving
# -----------------------------
def load_dashboard_html() -> str:
    if HTML_PATH.exists():
        return HTML_PATH.read_text(encoding="utf-8")
    return f"""
    <!doctype html>
    <html lang="pt-br">
    <head><meta charset="utf-8" /><title>Monitor Rastreasul</title></head>
    <body><h2>index.html não encontrado</h2><p>Arquivo esperado em: {HTML_PATH}</p></body>
    </html>
    """


def guess_content_type(file_path: Path) -> str:
    content_type, _ = mimetypes.guess_type(str(file_path))
    if content_type:
        if content_type.startswith("text/") or content_type in (
            "application/javascript",
            "application/json",
            "application/xml",
        ):
            return f"{content_type}; charset=utf-8"
        return content_type

    suffix = file_path.suffix.lower()
    if suffix == ".js":
        return "application/javascript; charset=utf-8"
    if suffix == ".css":
        return "text/css; charset=utf-8"
    if suffix == ".html":
        return "text/html; charset=utf-8"
    if suffix == ".json":
        return "application/json; charset=utf-8"
    if suffix == ".png":
        return "image/png"
    if suffix in (".jpg", ".jpeg"):
        return "image/jpeg"
    if suffix == ".svg":
        return "image/svg+xml"
    if suffix == ".ico":
        return "image/x-icon"
    return "application/octet-stream"


def resolve_static_path(request_path: str) -> Path | None:
    raw_path = unquote(urlparse(request_path).path)

    if raw_path in ("", "/"):
        return HTML_PATH

    frontend_root = FRONTEND_DIR.resolve()
    relative = raw_path.lstrip("/")

    if not relative:
        return HTML_PATH

    candidate = (FRONTEND_DIR / relative).resolve()

    try:
        candidate.relative_to(frontend_root)
    except ValueError:
        return None

    if candidate.is_file():
        return candidate

    if candidate.is_dir():
        index_candidate = (candidate / "index.html").resolve()
        try:
            index_candidate.relative_to(frontend_root)
        except ValueError:
            return None
        if index_candidate.exists() and index_candidate.is_file():
            return index_candidate

    html_candidate = (FRONTEND_DIR / f"{relative}.html").resolve()
    try:
        html_candidate.relative_to(frontend_root)
    except ValueError:
        return None

    if html_candidate.exists() and html_candidate.is_file():
        return html_candidate

    return None


def get_allowed_origin(request_handler) -> str:
    request_origin = request_handler.headers.get("Origin")
    if not request_origin:
        return "*"
    if request_origin in CORS_ALLOWED_ORIGINS:
        return request_origin
    return "null"


# -----------------------------
# Treatment helpers
# -----------------------------
REQUIRED_TREATMENT_SHEET = "Mensagens"
TRIPS_TREATMENT_SHEET = "Horas de motor"
NUMERIC_COLUMNS_TREATMENT = {
    "tensao", "bateria", "temperatura ambiente", "odometro",
    "pressao p1e1", "temperatura p1e1", "pressao p2e1", "temperatura p2e1",
    "pressao p1e2", "temperatura p1e2", "pressao p2e2", "temperatura p2e2",
    "pressao p3e2", "temperatura p3e2", "pressao p4e2", "temperatura p4e2",
    "pressao p1e3", "temperatura p1e3", "pressao p2e3", "temperatura p2e3",
    "rpm", "carga do motor", "temperatura do motor", "velocidade", "velocidade_2",
    "rpm max", "acelerador", "nivel de combustivel", "consumo instantaneo",
    "consumido", "arla", "volante", "inercia", "aceleracao", "desaceleracao",
    "peso dianteiro", "pesoraseiro", "pesootal", "altitude"
}

ROUND_2_COLUMNS_TREATMENT = {
    "tensao", "bateria", "temperatura ambiente", "velocidade", "velocidade_2",
    "rpm max", "carga do motor", "acelerador", "nivel de combustivel",
    "inercia", "aceleracao"
}

ANALYSIS_HEADERS_TREATMENT = [
    "dia-mês",
    "Faixa Verde",
    "Trânsito (Velocidade)",
    "Trânsito (RPM)",
    "Aceleração",
    "Frenagens",
    "Nota Rota",
    "Soma Pontuação",
    "Nota Motorista",
    "Viagem",
]

NORMALIZE_KEY_TRANSLATION_TREATMENT = str.maketrans({
    "á": "a", "à": "a", "â": "a", "ã": "a",
    "é": "e", "ê": "e",
    "í": "i",
    "ó": "o", "ô": "o", "õ": "o",
    "ú": "u",
    "ç": "c",
})
NUMBER_TOKEN_PATTERN_TREATMENT = re.compile(r"-?\d[\d.,]*")
PROGRESS_UPDATE_EVERY_TREATMENT = int(os.getenv("PROGRESS_UPDATE_EVERY_TREATMENT", "250"))
TREATMENT_OPEN_FROM_PATH = os.getenv("TREATMENT_OPEN_FROM_PATH", "1") == "1"
DATE_FORMATS_SLASH_WITH_TIME_4_TREATMENT = ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M")
DATE_FORMATS_SLASH_WITH_TIME_2_TREATMENT = ("%d/%m/%y %H:%M:%S", "%d/%m/%y %H:%M")
DATE_FORMATS_SLASH_DATE_4_TREATMENT = ("%d/%m/%Y",)
DATE_FORMATS_SLASH_DATE_2_TREATMENT = ("%d/%m/%y",)
DATE_FORMATS_DASH_WITH_TIME_TREATMENT = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M")
DATE_FORMATS_DASH_DATE_TREATMENT = ("%Y-%m-%d",)


def cleanup_old_treatment_jobs(max_age_seconds: int = 3600):
    now = time.time()
    to_delete = []

    with TREATMENT_JOBS_LOCK:
        for job_id, meta in list(TREATMENT_JOBS.items()):
            created_at = float(meta.get("created_at", 0))
            if now - created_at > max_age_seconds:
                to_delete.append((job_id, meta.get("path")))

        for job_id, _ in to_delete:
            TREATMENT_JOBS.pop(job_id, None)

    for _, path in to_delete:
        try:
            if path:
                Path(path).unlink(missing_ok=True)
        except Exception:
            pass

    status_delete = []
    with TREATMENT_STATUS_LOCK:
        for job_id, meta in list(TREATMENT_STATUS.items()):
            created_at = float(meta.get("created_at", 0))
            if now - created_at > max_age_seconds:
                status_delete.append(meta)
                TREATMENT_STATUS.pop(job_id, None)

    for meta in status_delete:
        for key in ("input_path",):
            try:
                path = meta.get(key)
                if path:
                    Path(path).unlink(missing_ok=True)
            except Exception:
                pass


def register_treatment_job(file_path: Path, download_name: str, job_id: str | None = None):
    cleanup_old_treatment_jobs()
    job_id = job_id or uuid.uuid4().hex
    with TREATMENT_JOBS_LOCK:
        TREATMENT_JOBS[job_id] = {
            "path": str(file_path),
            "download_name": download_name,
            "created_at": time.time(),
        }
    return job_id


def get_treatment_job(job_id: str):
    cleanup_old_treatment_jobs()
    with TREATMENT_JOBS_LOCK:
        return TREATMENT_JOBS.get(job_id)


def create_treatment_status(job_id: str, kind: str, filename: str, input_path: Path):
    cleanup_old_treatment_jobs()
    now = time.time()
    with TREATMENT_STATUS_LOCK:
        TREATMENT_STATUS[job_id] = {
            "job_id": job_id,
            "kind": kind,
            "status": "queued",
            "created_at": now,
            "updated_at": now,
            "filename": filename,
            "input_path": str(input_path),
            "progress": {
                "phase": "upload_received",
                "message": "Upload concluído. Preparando processamento...",
                "current": None,
                "total": None,
            },
            "result": None,
            "error": None,
        }


def update_treatment_status(job_id: str, **fields):
    with TREATMENT_STATUS_LOCK:
        current = TREATMENT_STATUS.get(job_id)
        if not current:
            return None
        current.update(fields)
        current["updated_at"] = time.time()
        return dict(current)


def update_treatment_progress(job_id: str, phase: str, message: str, current=None, total=None, status: str | None = None):
    payload = {
        "progress": {
            "phase": phase,
            "message": message,
            "current": current,
            "total": total,
        }
    }
    if status:
        payload["status"] = status
    return update_treatment_status(job_id, **payload)


def get_treatment_status(job_id: str):
    cleanup_old_treatment_jobs()
    with TREATMENT_STATUS_LOCK:
        status = TREATMENT_STATUS.get(job_id)
        return dict(status) if status else None


def save_treatment_input_file(file_bytes: bytes, filename: str, job_id: str):
    suffix = Path(filename or "arquivo.xlsx").suffix or ".xlsx"
    input_path = TEMP_TREATMENT_INBOX_DIR / f"{job_id}{suffix}"
    input_path.write_bytes(file_bytes)
    return input_path


def normalize_spaces_treatment(value):
    if value is None:
        return ""
    if isinstance(value, str):
        if not value:
            return ""
        text = value.replace("\u00A0", " ")
        stripped = text.strip()
        if not stripped:
            return ""
        if "  " not in stripped and "\t" not in stripped and "\n" not in stripped and "\r" not in stripped:
            return stripped
        return " ".join(stripped.split())
    return " ".join(str(value).replace("\u00A0", " ").split()).strip()


def normalize_header_treatment(value, fallback_index):
    text = normalize_spaces_treatment(value).replace("*", "")
    return text or f"Coluna_{fallback_index + 1}"


def normalize_key_treatment(value):
    return normalize_spaces_treatment(value).replace("*", "").lower().translate(NORMALIZE_KEY_TRANSLATION_TREATMENT)


def make_unique_headers_treatment(headers):
    seen = {}
    out = []
    for header in headers:
        count = seen.get(header, 0) + 1
        seen[header] = count
        out.append(header if count == 1 else f"{header}_{count}")
    return out


def is_row_empty_treatment(row):
    if not isinstance(row, (list, tuple)):
        return True
    for cell in row:
        if cell is None:
            continue
        if isinstance(cell, str):
            if normalize_spaces_treatment(cell) != "":
                return False
            continue
        return False
    return True


def worksheet_to_rows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def find_header_in_worksheet_treatment(ws):
    first_non_empty = None

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        row_values = list(row)
        if isinstance(row_values, list) and any(normalize_key_treatment(cell) == "hora" for cell in row_values):
            return idx, row_values

        if first_non_empty is None and not is_row_empty_treatment(row_values):
            first_non_empty = (idx, row_values)

    return first_non_empty or (-1, [])


def build_output_headers_treatment(headers, key_map):
    output_headers = []

    for header, key in zip(headers, key_map):
        if key == "agrupamento":
            continue

        output_headers.append(header)
        if key == "hora":
            output_headers.append("Hora Unix")

    output_headers.append("Linha Origem")
    return output_headers


def build_column_modes_treatment(key_map):
    modes = []
    for key in key_map:
        if key == "agrupamento":
            modes.append("skip")
        elif key == "hora":
            modes.append("hour")
        elif key in NUMERIC_COLUMNS_TREATMENT:
            modes.append("numeric")
        else:
            modes.append("text")
    return modes


def normalize_cell_for_treatment(raw_value, key, mode=None):
    mode = mode or key

    if mode == "hour":
        parsed = parse_brazil_datetime_treatment(raw_value)
        return (
            format_datetime_br_treatment(parsed) if parsed else normalize_spaces_treatment(raw_value),
            int(parsed.timestamp()) if parsed else "",
        )

    if mode == "numeric":
        numeric = to_number_maybe_treatment(raw_value)
        if isinstance(numeric, (int, float)):
            return round_if_needed_treatment(numeric, key)
        return numeric

    if isinstance(raw_value, datetime):
        return format_datetime_br_treatment(raw_value)

    return normalize_spaces_treatment(raw_value)


def transform_row_values_treatment(row, key_map, column_modes, source_line_number):
    output_values = []
    has_any_data = False
    row_len = len(row)
    append_value = output_values.append

    for col_index, mode in enumerate(column_modes):
        if mode == "skip":
            continue

        raw_value = row[col_index] if col_index < row_len else None
        key = key_map[col_index]

        if mode == "hour":
            hour_value, hour_unix = normalize_cell_for_treatment(raw_value, key, mode)
            append_value(hour_value)
            append_value(hour_unix)
            if hour_value != "" or hour_unix != "":
                has_any_data = True
            continue

        cell_value = normalize_cell_for_treatment(raw_value, key, mode)
        append_value(cell_value)
        if cell_value != "":
            has_any_data = True

    if not has_any_data:
        return None

    append_value(source_line_number)
    return output_values


def make_preview_row_treatment(headers, values):
    return {header: value for header, value in zip(headers, values)}


def prepare_treatment_stream(ws):
    header_index, header_row = find_header_in_worksheet_treatment(ws)
    if header_index < 0:
        raise ValueError("Não foi possível localizar o cabeçalho.")

    raw_headers = [normalize_header_treatment(value, idx) for idx, value in enumerate(header_row or [])]
    headers = make_unique_headers_treatment(raw_headers)
    key_map = [normalize_key_treatment(header) for header in headers]
    column_modes = build_column_modes_treatment(key_map)
    output_headers = build_output_headers_treatment(headers, key_map)

    return {
        "header_index": header_index,
        "key_map": key_map,
        "column_modes": column_modes,
        "output_headers": output_headers,
    }


def iter_processed_rows_treatment_from_worksheet(ws, stream_meta):
    header_index = stream_meta["header_index"]
    key_map = stream_meta["key_map"]
    column_modes = stream_meta["column_modes"]

    for row_index, row in enumerate(ws.iter_rows(values_only=True)):
        if row_index <= header_index:
            continue
        if is_row_empty_treatment(row):
            continue

        transformed = transform_row_values_treatment(
            row=row,
            key_map=key_map,
            column_modes=column_modes,
            source_line_number=row_index + 1,
        )
        if transformed is None:
            continue

        yield transformed


def excel_serial_to_datetime_treatment(value):
    if not isinstance(value, (int, float)):
        return None
    try:
        base = datetime(1899, 12, 30)
        return base + timedelta(days=float(value))
    except Exception:
        return None


def parse_brazil_datetime_treatment(value):
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, (int, float)):
        return excel_serial_to_datetime_treatment(value)

    text = normalize_spaces_treatment(value)
    if not text:
        return None

    if "/" in text:
        if " " in text:
            date_part = text.split(" ", 1)[0]
            year_len = len(date_part.rsplit("/", 1)[-1])
            formats = DATE_FORMATS_SLASH_WITH_TIME_4_TREATMENT if year_len == 4 else DATE_FORMATS_SLASH_WITH_TIME_2_TREATMENT
        else:
            year_len = len(text.rsplit("/", 1)[-1])
            formats = DATE_FORMATS_SLASH_DATE_4_TREATMENT if year_len == 4 else DATE_FORMATS_SLASH_DATE_2_TREATMENT
    elif "-" in text:
        formats = DATE_FORMATS_DASH_WITH_TIME_TREATMENT if " " in text else DATE_FORMATS_DASH_DATE_TREATMENT
    else:
        return None

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass

    return None


def format_datetime_br_treatment(value):
    if not isinstance(value, datetime):
        return ""
    return value.strftime("%d/%m/%Y %H:%M:%S")


def token_to_number_treatment(token):
    if token is None:
        return None

    value = str(token).replace("−", "-").strip()
    if not value:
        return None

    has_dot = "." in value
    has_comma = "," in value

    if has_dot and has_comma:
        if value.rfind(".") > value.rfind(","):
            value = value.replace(",", "")
        else:
            value = value.replace(".", "").replace(",", ".")
    elif has_comma:
        if value.count(",") > 1:
            last = value.rfind(",")
            value = value[:last].replace(",", "") + "." + value[last + 1:]
        else:
            value = value.replace(",", ".")
    elif has_dot:
        if value.count(".") > 1:
            last = value.rfind(".")
            value = value[:last].replace(".", "") + "." + value[last + 1:]

    try:
        num = float(value)
        return num
    except Exception:
        return None


def to_number_maybe_treatment(value):
    if value in (None, ""):
        return ""

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value

    original = normalize_spaces_treatment(value)
    if not original:
        return ""

    if any(ch.isdigit() for ch in original):
        simple_candidate = original.replace("−", "-").replace(" ", "")
        if all(ch.isdigit() or ch in "-.,+" for ch in simple_candidate):
            num = token_to_number_treatment(simple_candidate)
            return num if num is not None else original
    else:
        return original

    best_candidate = None
    best_score = (-1, -1)
    for match in NUMBER_TOKEN_PATTERN_TREATMENT.finditer(original):
        candidate = match.group(0)
        digit_count = sum(ch.isdigit() for ch in candidate)
        score = (digit_count, len(candidate))
        if score > best_score:
            best_candidate = candidate
            best_score = score

    if not best_candidate:
        return original

    num = token_to_number_treatment(best_candidate)
    return num if num is not None else original


def round_if_needed_treatment(value, key):
    if not isinstance(value, (int, float)):
        return value
    if key not in ROUND_2_COLUMNS_TREATMENT:
        return value
    return round(float(value), 2)


def find_header_index_treatment(sheet_rows):
    for idx, row in enumerate(sheet_rows):
        if isinstance(row, (list, tuple)) and any(normalize_key_treatment(cell) == "hora" for cell in row):
            return idx

    for idx, row in enumerate(sheet_rows):
        if isinstance(row, (list, tuple)) and not is_row_empty_treatment(row):
            return idx

    return -1


def process_sheet_rows_treatment(sheet_rows):
    if not sheet_rows:
        raise ValueError("A planilha não possui dados válidos.")

    header_index = find_header_index_treatment(sheet_rows)
    if header_index < 0:
        raise ValueError("Não foi possível localizar o cabeçalho.")

    header_row = sheet_rows[header_index] or []
    raw_headers = [normalize_header_treatment(value, idx) for idx, value in enumerate(header_row)]
    headers = make_unique_headers_treatment(raw_headers)
    key_map = [normalize_key_treatment(header) for header in headers]

    data_rows = [row for row in sheet_rows[header_index + 1:] if not is_row_empty_treatment(row)]
    treated_rows = []

    for row_index, row in enumerate(data_rows):
        obj = {}
        has_any_data = False

        for col_index, header in enumerate(headers):
            raw_value = row[col_index] if col_index < len(row) else None
            key = key_map[col_index]

            if key == "agrupamento":
                continue

            if key == "hora":
                parsed = parse_brazil_datetime_treatment(raw_value)
                obj[header] = format_datetime_br_treatment(parsed) if parsed else normalize_spaces_treatment(raw_value)
                obj["Hora Unix"] = int(parsed.timestamp()) if parsed else ""
                if obj[header] != "" or obj["Hora Unix"] != "":
                    has_any_data = True
                continue

            if key in NUMERIC_COLUMNS_TREATMENT:
                numeric = to_number_maybe_treatment(raw_value)
                obj[header] = round_if_needed_treatment(numeric, key) if isinstance(numeric, (int, float)) else numeric
                if obj[header] != "":
                    has_any_data = True
                continue

            if isinstance(raw_value, datetime):
                obj[header] = format_datetime_br_treatment(raw_value)
            else:
                obj[header] = normalize_spaces_treatment(raw_value)

            if obj[header] != "":
                has_any_data = True

        if not has_any_data:
            continue

        obj["Linha Origem"] = header_index + row_index + 2
        treated_rows.append(obj)

    if not treated_rows:
        raise ValueError("Nenhuma linha válida foi encontrada para tratamento.")

    return {
        "headers": list(treated_rows[0].keys()),
        "rows": treated_rows,
    }


def find_sheet_name_treatment(workbook, desired_name):
    if desired_name in workbook.sheetnames:
        return desired_name

    normalized_target = normalize_key_treatment(desired_name)
    for name in workbook.sheetnames:
        if normalize_key_treatment(name) == normalized_target:
            return name

    return None


def process_trips_sheet_rows_treatment(sheet_rows):
    if not sheet_rows:
        return {"headers": [], "rows": [], "trips": []}

    treated = process_sheet_rows_treatment(sheet_rows)
    headers = treated["headers"]

    start_header = next((h for h in headers if normalize_key_treatment(h) in {"inicio", "início"}), None)
    end_header = next((h for h in headers if normalize_key_treatment(h) == "fim"), None)

    if not start_header or not end_header:
        return {"headers": treated["headers"], "rows": treated["rows"], "trips": []}

    trips = []
    for row in treated["rows"]:
        start_date = parse_brazil_datetime_treatment(row.get(start_header))
        end_date = parse_brazil_datetime_treatment(row.get(end_header))
        if not start_date or not end_date:
            continue
        if end_date < start_date:
            continue
        trips.append({"start": start_date, "end": end_date})

    trips.sort(key=lambda item: item["start"])

    return {
        "headers": treated["headers"],
        "rows": treated["rows"],
        "trips": [
            {"id": f"Viagem {idx + 1}", "start": item["start"], "end": item["end"]}
            for idx, item in enumerate(trips)
        ],
    }


def get_preview_rows(rows):
    return rows[:PREVIEW_LIMIT]


def build_treatment_export_name(filename):
    if not filename:
        return "arquivo_tratado.xlsx"
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return filename[:-5] + "_tratado.xlsx"
    if lower.endswith(".xls"):
        return filename[:-4] + "_tratado.xlsx"
    return filename + "_tratado.xlsx"


def build_analysis_export_name(filename):
    if not filename:
        return "arquivo_tratado_analise.xlsx"
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return filename[:-5] + "_tratado_analise.xlsx"
    if lower.endswith(".xls"):
        return filename[:-4] + "_tratado_analise.xlsx"
    return filename + "_tratado_analise.xlsx"


def write_result_workbook(output_path: Path, sheets: list[tuple[str, list[str], list[dict]]]):
    wb = Workbook(write_only=True)
    first = True

    for sheet_name, headers, rows in sheets:
        ws = wb.create_sheet(title=(sheet_name or "Sheet1")[:31])
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])

        if first and "Sheet" in wb.sheetnames:
            default_ws = wb["Sheet"]
            wb.remove(default_ws)
            first = False

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(wb["Sheet"])

    wb.save(output_path)


def score_intensity_label_treatment(value):
    if value == "Leve":
        return 2
    if value in {"Médio", "Média"}:
        return 1
    if value in {"Intenso", "Intensa"}:
        return 0
    return 0


def score_green_band_treatment(value):
    return 2 if value == "Faixa Verde" else 0


def calculate_sum_score_treatment(green_band, speed_transit, rpm_transit, acceleration, brakes):
    return (
        score_intensity_label_treatment(rpm_transit)
        + score_intensity_label_treatment(acceleration)
        + score_intensity_label_treatment(brakes)
        + score_green_band_treatment(green_band)
        + score_intensity_label_treatment(speed_transit)
    )


def get_route_score_treatment(value):
    if value == "Intenso":
        return 10
    if value in {"Médio", "Média"}:
        return 6
    if value == "Leve":
        return 0
    return ""


def calculate_driver_score_treatment(sum_score, route_score):
    if not isinstance(sum_score, (int, float)):
        return ""
    if not isinstance(route_score, (int, float)) or route_score == 0:
        return "100%"
    return f"{round((sum_score / route_score) * 100)}%"


def classify_green_band_treatment(rpm_value, config):
    if not isinstance(rpm_value, (int, float)):
        return ""
    return "Faixa Verde" if config["bestRpmMin"] <= rpm_value <= config["bestRpmMax"] else "Fora da Faixa"


def classify_speed_transit_treatment(speed_value, config):
    if not isinstance(speed_value, (int, float)):
        return ""
    if speed_value <= config["speedLowMax"]:
        return "Intenso"
    if speed_value <= config["speedMediumMax"]:
        return "Médio"
    return "Leve"


def classify_rpm_transit_treatment(rpm_value, config):
    if not isinstance(rpm_value, (int, float)):
        return ""
    if rpm_value < config["rpmUsefulStart"] or rpm_value > config["rpmUsefulEnd"]:
        return "Intenso"
    if rpm_value < config["rpmLightStart"]:
        return "Médio"
    return "Leve"


def classify_acceleration_treatment(accel_value, config):
    if not isinstance(accel_value, (int, float)):
        return ""
    if accel_value <= config["accelLightMax"]:
        return "Leve"
    if accel_value <= config["accelMediumMax"]:
        return "Média"
    return "Intensa"


def classify_brake_treatment(brake_count, config):
    if not isinstance(brake_count, (int, float)):
        return ""
    if brake_count < config["brakeMediumMin"]:
        return "Leve"
    if brake_count < config["brakeIntenseMin"]:
        return "Média"
    return "Intensa"


def is_brake_activated_treatment(value):
    if value in (None, ""):
        return False
    if isinstance(value, (int, float)):
        return value > 0
    normalized = normalize_key_treatment(value)
    if normalized in {"sim", "true", "on", "ativo", "acionado", "pressed"}:
        return True
    if normalized in {"nao", "não", "false", "off", "inativo", "desacionado"}:
        return False
    numeric = token_to_number_treatment(normalized)
    return (numeric or 0) > 0 if numeric is not None else False


def find_header_by_aliases_treatment(headers, aliases):
    aliases_norm = {normalize_key_treatment(alias) for alias in aliases}
    for header in headers:
        if normalize_key_treatment(header) in aliases_norm:
            return header
    return None


def extract_numeric_from_row_treatment(row, header_names):
    for header_name in header_names:
        if not header_name:
            continue
        parsed = to_number_maybe_treatment(row.get(header_name))
        if isinstance(parsed, (int, float)):
            return parsed
    return None


def get_row_date_treatment(row, hour_header, hour_unix_header):
    if hour_unix_header:
        unix = extract_numeric_from_row_treatment(row, [hour_unix_header])
        if isinstance(unix, (int, float)):
            try:
                return datetime.fromtimestamp(unix)
            except Exception:
                pass

    if hour_header:
        return parse_brazil_datetime_treatment(row.get(hour_header))

    return None


def format_day_month_treatment(date):
    if not isinstance(date, datetime):
        return ""
    return date.strftime("%d/%m")


def find_trip_id_treatment(row_date, trips):
    if not isinstance(row_date, datetime) or not trips:
        return ""

    row_ts = row_date.timestamp()
    trip_starts = [trip["start"].timestamp() for trip in trips]
    idx = bisect_left(trip_starts, row_ts)
    candidates = []

    if idx < len(trips):
        candidates.append(trips[idx])
    if idx > 0:
        candidates.append(trips[idx - 1])

    for trip in candidates:
        if trip["start"].timestamp() <= row_ts <= trip["end"].timestamp():
            return trip["id"]
    return ""


def resolve_header_index_treatment(headers, aliases):
    header = find_header_by_aliases_treatment(headers, aliases)
    if not header:
        return None
    try:
        return headers.index(header)
    except ValueError:
        return None


def get_numeric_from_indexed_row_treatment(values, idx):
    if idx is None or idx >= len(values):
        return None
    parsed = to_number_maybe_treatment(values[idx])
    return parsed if isinstance(parsed, (int, float)) else None


def get_datetime_from_indexed_row_treatment(values, hour_idx, hour_unix_idx):
    if hour_unix_idx is not None and hour_unix_idx < len(values):
        unix = get_numeric_from_indexed_row_treatment(values, hour_unix_idx)
        if isinstance(unix, (int, float)):
            try:
                return datetime.fromtimestamp(unix)
            except Exception:
                pass

    if hour_idx is not None and hour_idx < len(values):
        return parse_brazil_datetime_treatment(values[hour_idx])

    return None


def build_trip_search_index_treatment(trips):
    indexed_trips = []

    for trip in trips or []:
        start = trip.get("start")
        end = trip.get("end")
        if not isinstance(start, datetime) or not isinstance(end, datetime):
            continue
        indexed_trips.append({
            "id": trip.get("id", ""),
            "start_ts": start.timestamp(),
            "end_ts": end.timestamp(),
        })

    trip_starts = [trip["start_ts"] for trip in indexed_trips]
    return indexed_trips, trip_starts


def find_trip_id_with_index_treatment(row_date, indexed_trips, trip_starts):
    if not isinstance(row_date, datetime) or not indexed_trips:
        return ""

    row_ts = row_date.timestamp()
    idx = bisect_left(trip_starts, row_ts)

    if idx < len(indexed_trips):
        trip = indexed_trips[idx]
        if trip["start_ts"] <= row_ts <= trip["end_ts"]:
            return trip["id"]

    if idx > 0:
        trip = indexed_trips[idx - 1]
        if trip["start_ts"] <= row_ts <= trip["end_ts"]:
            return trip["id"]

    return ""


def build_analysis_rows_treatment(rows, headers, trips, config):
    hour_header = find_header_by_aliases_treatment(headers, ["Hora"])
    hour_unix_header = find_header_by_aliases_treatment(headers, ["Hora Unix", "HoraUnix"])
    rpm_header = find_header_by_aliases_treatment(headers, ["RPM"])
    speed_header = find_header_by_aliases_treatment(headers, ["Velocidade_2", "Velocidade"])
    accel_header = find_header_by_aliases_treatment(headers, ["Acelerador"])
    brake_header = find_header_by_aliases_treatment(headers, ["Freio"])

    recent_brake_events = []
    previous_brake_active = False
    analysis_rows = []

    for row in rows:
        current_date = get_row_date_treatment(row, hour_header, hour_unix_header)
        current_ts = current_date.timestamp() if current_date else None
        rpm_value = extract_numeric_from_row_treatment(row, [rpm_header])
        speed_value = extract_numeric_from_row_treatment(row, [speed_header])
        accel_value = extract_numeric_from_row_treatment(row, [accel_header])
        brake_active = is_brake_activated_treatment(row.get(brake_header)) if brake_header else False

        if current_ts is not None:
            recent_brake_events = [x for x in recent_brake_events if x >= current_ts - 60]
            if brake_active and not previous_brake_active:
                recent_brake_events.append(current_ts)

        speed_transit = classify_speed_transit_treatment(speed_value, config)
        note_route = get_route_score_treatment(speed_transit)
        brake_count = len(recent_brake_events) if current_ts is not None and brake_header else None
        brake_label = classify_brake_treatment(brake_count, config)
        green_band = classify_green_band_treatment(rpm_value, config)
        rpm_transit = classify_rpm_transit_treatment(rpm_value, config)
        acceleration_label = classify_acceleration_treatment(accel_value, config)
        sum_score = calculate_sum_score_treatment(
            green_band=green_band,
            speed_transit=speed_transit,
            rpm_transit=rpm_transit,
            acceleration=acceleration_label,
            brakes=brake_label,
        )
        driver_score = calculate_driver_score_treatment(sum_score, note_route)
        trip_id = find_trip_id_treatment(current_date, trips)

        previous_brake_active = brake_active

        merged = dict(row)
        merged.update({
            "dia-mês": format_day_month_treatment(current_date),
            "Faixa Verde": green_band,
            "Trânsito (Velocidade)": speed_transit,
            "Trânsito (RPM)": rpm_transit,
            "Aceleração": acceleration_label,
            "Frenagens": brake_label,
            "Nota Rota": note_route,
            "Soma Pontuação": sum_score,
            "Nota Motorista": driver_score,
            "Viagem": trip_id,
        })
        analysis_rows.append(merged)

    base_headers = [h for h in headers if h not in ANALYSIS_HEADERS_TREATMENT]
    analysis_headers = base_headers + ANALYSIS_HEADERS_TREATMENT
    return analysis_headers, analysis_rows


def open_treatment_workbook(file_bytes: bytes | None = None, input_path: Path | None = None):
    if TREATMENT_OPEN_FROM_PATH and input_path:
        return load_workbook(input_path, data_only=True, read_only=True), "path"
    if file_bytes is None:
        raise ValueError("file_bytes ausente para abertura do workbook.")
    return load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True), "bytes"


def process_treatment_file_bytes(
    file_bytes: bytes | None,
    filename: str,
    job_id: str | None = None,
    progress_cb=None,
    input_path: Path | None = None,
):
    started_at = time.perf_counter()
    if progress_cb:
        progress_cb("opening_workbook", "Abrindo planilha...")
    wb, open_mode = open_treatment_workbook(file_bytes=file_bytes, input_path=input_path)
    load_completed_at = time.perf_counter()
    sheet_name = find_sheet_name_treatment(wb, REQUIRED_TREATMENT_SHEET)
    sheet_resolved_at = time.perf_counter()
    if not sheet_name:
        raise ValueError('A planilha obrigatória "Mensagens" não foi encontrada no arquivo.')

    ws = wb[sheet_name]
    stream_meta = prepare_treatment_stream(ws)
    stream_prepared_at = time.perf_counter()
    headers = stream_meta["output_headers"]

    export_name = build_treatment_export_name(filename)
    output_path = TEMP_TREATMENT_DIR / f"{uuid.uuid4().hex}_{export_name}"
    output_wb = Workbook(write_only=True)
    output_ws = output_wb.create_sheet(title="Tratado")
    output_ws.append(headers)

    preview_rows = []
    row_count = 0
    if progress_cb:
        progress_cb("processing_rows", "Tratando linhas da planilha...")
    for row_values in iter_processed_rows_treatment_from_worksheet(ws, stream_meta):
        output_ws.append(row_values)
        row_count += 1
        if len(preview_rows) < PREVIEW_LIMIT:
            preview_rows.append(make_preview_row_treatment(headers, row_values))
        if progress_cb and row_count % PROGRESS_UPDATE_EVERY_TREATMENT == 0:
            progress_cb("processing_rows", "Tratando linhas da planilha...", current=row_count, total=None)

    if row_count == 0:
        raise ValueError("Nenhuma linha válida foi encontrada para tratamento.")

    loop_completed_at = time.perf_counter()
    if "Sheet" in output_wb.sheetnames:
        output_wb.remove(output_wb["Sheet"])
    if progress_cb:
        progress_cb("writing_output", "Gerando arquivo final para download...", current=row_count, total=None)
    output_wb.save(output_path)
    save_completed_at = time.perf_counter()

    job_id = register_treatment_job(output_path, export_name, job_id=job_id)
    finished_at = time.perf_counter()
    total_seconds = finished_at - started_at
    rows_per_second = (row_count / total_seconds) if total_seconds > 0 else 0.0
    rows_per_minute = rows_per_second * 60.0

    print(
        f"[treatment][job_id={job_id or '-'}][mode={open_mode}] "
        "process_treatment "
        f"total={total_seconds:.2f}s "
        f"load={load_completed_at - started_at:.2f}s "
        f"sheet={sheet_resolved_at - load_completed_at:.2f}s "
        f"prepare={stream_prepared_at - sheet_resolved_at:.2f}s "
        f"process={loop_completed_at - stream_prepared_at:.2f}s "
        f"save={save_completed_at - loop_completed_at:.2f}s "
        f"register={finished_at - save_completed_at:.2f}s "
        f"rows={row_count} "
        f"rps={rows_per_second:.2f} "
        f"rpm={rows_per_minute:.2f}",
        flush=True,
    )

    return {
        "ok": True,
        "job_id": job_id,
        "original_file_name": filename,
        "export_file_name": export_name,
        "sheet_name": sheet_name,
        "row_count": row_count,
        "column_count": len(headers),
        "preview_headers": headers,
        "preview_rows": preview_rows,
    }


def process_treatment_step1_file_bytes(
    file_bytes: bytes | None,
    filename: str,
    config: dict,
    job_id: str | None = None,
    progress_cb=None,
    input_path: Path | None = None,
):
    started_at = time.perf_counter()
    if progress_cb:
        progress_cb("opening_workbook", "Abrindo planilha...")
    wb, open_mode = open_treatment_workbook(file_bytes=file_bytes, input_path=input_path)
    load_completed_at = time.perf_counter()

    target_sheet_name = find_sheet_name_treatment(wb, REQUIRED_TREATMENT_SHEET)
    target_sheet_resolved_at = time.perf_counter()
    if not target_sheet_name:
        raise ValueError('A planilha obrigatória "Mensagens" não foi encontrada no arquivo.')

    trips_sheet_name = find_sheet_name_treatment(wb, TRIPS_TREATMENT_SHEET)
    trips_headers = []
    trips_rows = []
    trips = []

    if trips_sheet_name:
        if progress_cb:
            progress_cb("processing_rows", "Lendo aba Horas de motor...")
        treated_trips = process_trips_sheet_rows_treatment(worksheet_to_rows(wb[trips_sheet_name]))
        trips_headers = treated_trips["headers"]
        trips_rows = treated_trips["rows"]
        trips = treated_trips["trips"]
    trips_processed_at = time.perf_counter()

    ws = wb[target_sheet_name]
    stream_meta = prepare_treatment_stream(ws)
    stream_prepared_at = time.perf_counter()
    treated_headers = stream_meta["output_headers"]
    analysis_headers = [header for header in treated_headers if header not in ANALYSIS_HEADERS_TREATMENT] + ANALYSIS_HEADERS_TREATMENT
    indexed_trips, trip_starts = build_trip_search_index_treatment(trips)
    analysis_context_ready_at = time.perf_counter()

    export_name = build_analysis_export_name(filename)
    output_path = TEMP_TREATMENT_DIR / f"{uuid.uuid4().hex}_{export_name}"
    output_wb = Workbook(write_only=True)
    analysis_ws = output_wb.create_sheet(title="Tratado Analise")
    analysis_ws.append(analysis_headers)

    hour_idx = resolve_header_index_treatment(treated_headers, ["Hora"])
    hour_unix_idx = resolve_header_index_treatment(treated_headers, ["Hora Unix", "HoraUnix"])
    rpm_idx = resolve_header_index_treatment(treated_headers, ["RPM"])
    speed_idx = resolve_header_index_treatment(treated_headers, ["Velocidade_2", "Velocidade"])
    accel_idx = resolve_header_index_treatment(treated_headers, ["Acelerador"])
    brake_idx = resolve_header_index_treatment(treated_headers, ["Freio"])

    recent_brake_events = []
    previous_brake_active = False
    preview_rows = []
    row_count = 0

    if progress_cb:
        progress_cb("processing_rows", "Tratando linhas da planilha com análise...")
    for treated_values in iter_processed_rows_treatment_from_worksheet(ws, stream_meta):
        current_date = get_datetime_from_indexed_row_treatment(
            treated_values,
            hour_idx,
            hour_unix_idx,
        )
        current_ts = current_date.timestamp() if current_date else None
        rpm_value = get_numeric_from_indexed_row_treatment(treated_values, rpm_idx)
        speed_value = get_numeric_from_indexed_row_treatment(treated_values, speed_idx)
        accel_value = get_numeric_from_indexed_row_treatment(treated_values, accel_idx)
        brake_active = (
            is_brake_activated_treatment(treated_values[brake_idx])
            if brake_idx is not None and brake_idx < len(treated_values)
            else False
        )

        if current_ts is not None:
            recent_brake_events = [ts for ts in recent_brake_events if ts >= current_ts - 60]
            if brake_active and not previous_brake_active:
                recent_brake_events.append(current_ts)

        speed_transit = classify_speed_transit_treatment(speed_value, config)
        note_route = get_route_score_treatment(speed_transit)
        brake_count = len(recent_brake_events) if current_ts is not None and brake_idx is not None else None
        brake_label = classify_brake_treatment(brake_count, config)
        green_band = classify_green_band_treatment(rpm_value, config)
        rpm_transit = classify_rpm_transit_treatment(rpm_value, config)
        acceleration_label = classify_acceleration_treatment(accel_value, config)
        sum_score = calculate_sum_score_treatment(
            green_band=green_band,
            speed_transit=speed_transit,
            rpm_transit=rpm_transit,
            acceleration=acceleration_label,
            brakes=brake_label,
        )
        driver_score = calculate_driver_score_treatment(sum_score, note_route)
        trip_id = find_trip_id_with_index_treatment(current_date, indexed_trips, trip_starts)
        previous_brake_active = brake_active

        analysis_values = [
            format_day_month_treatment(current_date),
            green_band,
            speed_transit,
            rpm_transit,
            acceleration_label,
            brake_label,
            note_route,
            sum_score,
            driver_score,
            trip_id,
        ]
        output_values = treated_values + analysis_values
        analysis_ws.append(output_values)
        row_count += 1

        if len(preview_rows) < PREVIEW_LIMIT:
            preview_rows.append(make_preview_row_treatment(analysis_headers, output_values))
        if progress_cb and row_count % PROGRESS_UPDATE_EVERY_TREATMENT == 0:
            progress_cb("processing_rows", "Tratando linhas da planilha com análise...", current=row_count, total=None)

    if row_count == 0:
        raise ValueError("Nenhuma linha válida foi encontrada para tratamento.")

    loop_completed_at = time.perf_counter()
    if trips_sheet_name and trips_headers and trips_rows:
        trips_ws = output_wb.create_sheet(title=(trips_sheet_name or "Sheet1")[:31])
        trips_ws.append(trips_headers)
        for row in trips_rows:
            trips_ws.append([row.get(header, "") for header in trips_headers])

    if "Sheet" in output_wb.sheetnames:
        output_wb.remove(output_wb["Sheet"])
    if progress_cb:
        progress_cb("writing_output", "Gerando arquivo final para download...", current=row_count, total=None)
    output_wb.save(output_path)
    save_completed_at = time.perf_counter()

    job_id = register_treatment_job(output_path, export_name, job_id=job_id)
    finished_at = time.perf_counter()
    total_seconds = finished_at - started_at
    rows_per_second = (row_count / total_seconds) if total_seconds > 0 else 0.0
    rows_per_minute = rows_per_second * 60.0

    print(
        f"[treatment][job_id={job_id or '-'}][mode={open_mode}] "
        "process_treatment_step1 "
        f"total={total_seconds:.2f}s "
        f"load={load_completed_at - started_at:.2f}s "
        f"sheet={target_sheet_resolved_at - load_completed_at:.2f}s "
        f"trips={trips_processed_at - target_sheet_resolved_at:.2f}s "
        f"prepare={stream_prepared_at - trips_processed_at:.2f}s "
        f"context={analysis_context_ready_at - stream_prepared_at:.2f}s "
        f"process={loop_completed_at - analysis_context_ready_at:.2f}s "
        f"save={save_completed_at - loop_completed_at:.2f}s "
        f"register={finished_at - save_completed_at:.2f}s "
        f"rows={row_count} "
        f"rps={rows_per_second:.2f} "
        f"rpm={rows_per_minute:.2f} "
        f"trips_count={len(trips)}",
        flush=True,
    )

    return {
        "ok": True,
        "job_id": job_id,
        "original_file_name": filename,
        "export_file_name": export_name,
        "sheet_name": target_sheet_name,
        "trips_count": len(trips),
        "row_count": row_count,
        "column_count": len(analysis_headers),
        "preview_headers": analysis_headers,
        "preview_rows": preview_rows,
    }


def run_treatment_job(job_id: str, temp_input_path: Path, filename: str):
    def progress_cb(phase, message, current=None, total=None):
        update_treatment_progress(job_id, phase, message, current=current, total=total, status="processing")

    try:
        update_treatment_progress(job_id, "opening_workbook", "Abrindo planilha...", status="processing")
        file_bytes = None if TREATMENT_OPEN_FROM_PATH else temp_input_path.read_bytes()
        result = process_treatment_file_bytes(
            file_bytes,
            filename,
            job_id=job_id,
            progress_cb=progress_cb,
            input_path=temp_input_path,
        )
        result["job_id"] = job_id
        update_treatment_status(
            job_id,
            status="done",
            result=result,
            error=None,
            progress={
                "phase": "done",
                "message": "Arquivo processado com sucesso",
                "current": None,
                "total": None,
            },
        )
    except Exception as e:
        update_treatment_status(
            job_id,
            status="error",
            result=None,
            error=str(e),
            progress={
                "phase": "error",
                "message": "Falha no processamento",
                "current": None,
                "total": None,
            },
        )
    finally:
        try:
            temp_input_path.unlink(missing_ok=True)
        except Exception:
            pass


def run_treatment_step1_job(job_id: str, temp_input_path: Path, filename: str, config: dict):
    def progress_cb(phase, message, current=None, total=None):
        update_treatment_progress(job_id, phase, message, current=current, total=total, status="processing")

    try:
        update_treatment_progress(job_id, "opening_workbook", "Abrindo planilha...", status="processing")
        file_bytes = None if TREATMENT_OPEN_FROM_PATH else temp_input_path.read_bytes()
        result = process_treatment_step1_file_bytes(
            file_bytes,
            filename,
            config,
            job_id=job_id,
            progress_cb=progress_cb,
            input_path=temp_input_path,
        )
        result["job_id"] = job_id
        update_treatment_status(
            job_id,
            status="done",
            result=result,
            error=None,
            progress={
                "phase": "done",
                "message": "Arquivo processado com sucesso",
                "current": None,
                "total": None,
            },
        )
    except Exception as e:
        update_treatment_status(
            job_id,
            status="error",
            result=None,
            error=str(e),
            progress={
                "phase": "error",
                "message": "Falha no processamento",
                "current": None,
                "total": None,
            },
        )
    finally:
        try:
            temp_input_path.unlink(missing_ok=True)
        except Exception:
            pass


# -----------------------------
# Wialon calls
# -----------------------------
def wialon_ajax(svc: str, params_obj: dict):
    url = f"{WIALON_HOST}/wialon/ajax.html"
    payload = {
        "svc": svc,
        "sid": get_sid(),
        "params": json.dumps(params_obj, ensure_ascii=False),
    }
    r = requests.post(url, data=payload, timeout=TIMEOUT_SECONDS)
    r.raise_for_status()
    return r.json()


def wialon_avl_evts(tm_cursor: int):
    url = f"{WIALON_HOST}/avl_evts"
    payload = {
        "sid": get_sid(),
        "params": json.dumps({"tm": int(tm_cursor)}, ensure_ascii=False),
    }
    r = requests.post(url, data=payload, timeout=TIMEOUT_SECONDS)
    r.raise_for_status()
    return r.json()


def build_params_obj():
    return {"itemIds": [get_item_id()]}


# -----------------------------
# calc_last parsing
# -----------------------------
def pick_sensor(resp, idx: int):
    resp = normalize_resp(resp)
    sensors = resp.get("sensors")
    if not isinstance(sensors, dict):
        return None
    item = sensors.get(str(idx))
    if not isinstance(item, dict):
        return None
    fmt = item.get("format")
    return {
        "raw": item.get("value"),
        "display": fmt.get("value") if isinstance(fmt, dict) else None,
    }


def pick_pos_metric(resp, key: str):
    resp = normalize_resp(resp)
    pos = resp.get("pos")
    if not isinstance(pos, dict):
        return None
    item = pos.get(key)
    if not isinstance(item, dict):
        return None
    fmt = item.get("format")
    return {
        "raw": item.get("value"),
        "display": fmt.get("value") if isinstance(fmt, dict) else None,
    }


def extract_calc_last_tm(resp):
    r = normalize_resp(resp)

    pos = r.get("pos")
    if isinstance(pos, dict) and isinstance(pos.get("t"), (int, float)):
        return int(pos["t"])

    if isinstance(r.get("tm"), (int, float)):
        return int(r["tm"])

    return None


def build_status_from_calc_last(calc_last_resp, source="calc_last", chosen_reason="bootstrap"):
    r = normalize_resp(calc_last_resp)
    fields = {
        "velocidade": pick_pos_metric(r, "s"),
        "altitude": pick_pos_metric(r, "z"),
        "rpm": pick_sensor(r, SENSOR_IDX["rpm"]),
        "consumido": pick_sensor(r, SENSOR_IDX["consumido"]),
        "pct_acelerado": pick_sensor(r, SENSOR_IDX["pct_acelerado"]),
        "temperatura_motor": pick_sensor(r, SENSOR_IDX["temperatura_motor"]),
        "ar_cond": pick_sensor(r, SENSOR_IDX["ar_cond"]),
        "freio": pick_sensor(r, 40),
        "arla": pick_sensor(r, 41),
        "consumido_delta": pick_sensor(r, 42),
        "peso_total": pick_sensor(r, 43),
        "motor": pick_sensor(r, 39),
    }
    return {
        "tm": r.get("tm"),
        "pos_tm": extract_calc_last_tm(r),
        "msg_tm": extract_calc_last_tm(r),
        "item_id": get_item_id(),
        "fields": fields,
        "source": source,
        "chosen_reason": chosen_reason,
    }


def values_from_status(status: dict):
    f = status.get("fields") or {}
    return {
        "velocidade": to_num((f.get("velocidade") or {}).get("raw")),
        "altitude": to_num((f.get("altitude") or {}).get("raw")),
        "rpm": to_num((f.get("rpm") or {}).get("raw")),
        "consumido": to_num((f.get("consumido") or {}).get("raw")),
        "pct_acelerado": to_num((f.get("pct_acelerado") or {}).get("raw")),
        "motor": to_num((f.get("motor") or {}).get("raw")),
        "temperatura_motor": to_num((f.get("temperatura_motor") or {}).get("raw")),
        "ar_cond": to_num((f.get("ar_cond") or {}).get("raw")),
        "freio": to_num((f.get("freio") or {}).get("raw")),
        "arla": to_num((f.get("arla") or {}).get("raw")),
        "consumido_delta": to_num((f.get("consumido_delta") or {}).get("raw")),
        "peso_total": to_num((f.get("peso_total") or {}).get("raw")),
    }


# -----------------------------
# avl_evts parsing
# -----------------------------
def get_sensor_value_from_map(sensor_map: dict, idx: int):
    if not isinstance(sensor_map, dict):
        return None
    item = sensor_map.get(str(idx))
    if item is None:
        item = sensor_map.get(idx)
    if isinstance(item, dict):
        if "value" in item:
            return item.get("value")
        if "raw" in item:
            return item.get("raw")
        to = item.get("to")
        if isinstance(to, dict) and "v" in to:
            return to.get("v")
    return None


def get_sensor_display_from_map(sensor_map: dict, idx: int):
    if not isinstance(sensor_map, dict):
        return None
    item = sensor_map.get(str(idx))
    if item is None:
        item = sensor_map.get(idx)
    if isinstance(item, dict):
        fmt = item.get("format")
        if isinstance(fmt, dict):
            return fmt.get("value")
        return item.get("display")
    return None


def extract_units_update_for_item(ev_resp: dict, item_id: int):
    units_update = ev_resp.get("units_update") or ev_resp.get("unitsUpdate")
    if not isinstance(units_update, dict):
        return None
    block = units_update.get(str(item_id))
    if block is None:
        block = units_update.get(item_id)
    return block if isinstance(block, dict) else None


def extract_global_sensors_map(ev_resp: dict, item_id: int):
    sensors = ev_resp.get("sensors")
    if not isinstance(sensors, dict):
        return {}
    block = sensors.get(str(item_id))
    if block is None:
        block = sensors.get(item_id)
    return block if isinstance(block, dict) else {}


def extract_best_event_for_item(ev_resp: dict, item_id: int):
    events = ev_resp.get("events")
    if not isinstance(events, list):
        return None

    best = None
    for e in events:
        if not isinstance(e, dict):
            continue
        if e.get("i") != item_id:
            continue
        d = e.get("d")
        if not isinstance(d, dict):
            continue
        t = d.get("t")
        if not isinstance(t, (int, float)):
            continue
        if best is None or int(t) > int(best.get("d", {}).get("t", 0)):
            best = e
    return best


def build_status_from_avl(ev_resp: dict, item_id: int):
    unit_update = extract_units_update_for_item(ev_resp, item_id) or {}
    unit_update_sensors = unit_update.get("sensors") if isinstance(unit_update.get("sensors"), dict) else {}
    global_sensors = extract_global_sensors_map(ev_resp, item_id)
    event = extract_best_event_for_item(ev_resp, item_id)

    d = event.get("d") if isinstance(event, dict) else {}
    d = d if isinstance(d, dict) else {}
    pos = d.get("pos") if isinstance(d.get("pos"), dict) else {}
    p = d.get("p") if isinstance(d.get("p"), dict) else {}

    msg_tm = d.get("t")
    if not isinstance(msg_tm, (int, float)):
        msg_tm = None
    else:
        msg_tm = int(msg_tm)

    velocidade = first_not_none(
        to_num(pos.get("s")),
        to_num(get_sensor_value_from_map(global_sensors, SENSOR_IDX["velocidade"])),
        to_num(get_sensor_value_from_map(unit_update_sensors, SENSOR_IDX["velocidade"])),
    )

    altitude = first_not_none(
        to_num(pos.get("z")),
        to_num(get_sensor_value_from_map(global_sensors, SENSOR_IDX["altitude"])),
        to_num(get_sensor_value_from_map(unit_update_sensors, SENSOR_IDX["altitude"])),
    )

    rpm = first_not_none(
        to_num(get_sensor_value_from_map(global_sensors, SENSOR_IDX["rpm"])),
        to_num(get_sensor_value_from_map(unit_update_sensors, SENSOR_IDX["rpm"])),
    )

    consumido = first_not_none(
        to_num(get_sensor_value_from_map(global_sensors, SENSOR_IDX["consumido"])),
        to_num(get_sensor_value_from_map(unit_update_sensors, SENSOR_IDX["consumido"])),
    )

    pct_acelerado = first_not_none(
        to_num(get_sensor_value_from_map(global_sensors, SENSOR_IDX["pct_acelerado"])),
        to_num(get_sensor_value_from_map(unit_update_sensors, SENSOR_IDX["pct_acelerado"])),
        to_num(p.get("can_acc_pedal")),
    )

    temperatura_motor = first_not_none(
        to_num(get_sensor_value_from_map(global_sensors, SENSOR_IDX["temperatura_motor"])),
        to_num(get_sensor_value_from_map(unit_update_sensors, SENSOR_IDX["temperatura_motor"])),
        to_num(p.get("can_coolant_temp")),
    )

    ar_cond = normalize_bool_like(
        first_not_none(
            get_sensor_value_from_map(global_sensors, SENSOR_IDX["ar_cond"]),
            get_sensor_value_from_map(unit_update_sensors, SENSOR_IDX["ar_cond"]),
        )
    )

    freio = normalize_bool_like(p.get("can_breaks"))

    fields = {
        "velocidade": sensor_field(velocidade),
        "altitude": sensor_field(altitude),
        "rpm": sensor_field(
            rpm,
            first_not_none(
                get_sensor_display_from_map(global_sensors, SENSOR_IDX["rpm"]),
                get_sensor_display_from_map(unit_update_sensors, SENSOR_IDX["rpm"]),
            ),
        ),
        "consumido": sensor_field(
            consumido,
            first_not_none(
                get_sensor_display_from_map(global_sensors, SENSOR_IDX["consumido"]),
                get_sensor_display_from_map(unit_update_sensors, SENSOR_IDX["consumido"]),
            ),
        ),
        "pct_acelerado": sensor_field(
            pct_acelerado,
            first_not_none(
                get_sensor_display_from_map(global_sensors, SENSOR_IDX["pct_acelerado"]),
                get_sensor_display_from_map(unit_update_sensors, SENSOR_IDX["pct_acelerado"]),
            ),
        ),
        "temperatura_motor": sensor_field(
            temperatura_motor,
            first_not_none(
                get_sensor_display_from_map(global_sensors, SENSOR_IDX["temperatura_motor"]),
                get_sensor_display_from_map(unit_update_sensors, SENSOR_IDX["temperatura_motor"]),
            ),
        ),
        "ar_cond": sensor_field(
            ar_cond,
            first_not_none(
                get_sensor_display_from_map(global_sensors, SENSOR_IDX["ar_cond"]),
                get_sensor_display_from_map(unit_update_sensors, SENSOR_IDX["ar_cond"]),
            ),
        ),
        "freio": sensor_field(freio),
        "arla": sensor_field(to_num(p.get("can_adblue_level"))),
        "consumido_delta": sensor_field(None),
        "peso_total": sensor_field(
            first_not_none(
                to_num(p.get("can_gross_comb_weight")),
                to_num(p.get("can_weight")),
            )
        ),
        "motor": sensor_field(normalize_bool_like(p.get("can_engine_turned_on"))),
    }

    useful = any((f or {}).get("raw") is not None for f in fields.values()) or msg_tm is not None

    return {
        "tm": msg_tm,
        "pos_tm": msg_tm,
        "msg_tm": msg_tm,
        "item_id": item_id,
        "fields": fields,
        "source": "avl_evts",
        "chosen_reason": "avl_match" if useful else "avl_no_useful_unit_data",
        "useful": useful,
    }


# -----------------------------
# Sample / status builders
# -----------------------------
def build_values_signature(values: dict):
    return (
        values.get("velocidade"),
        values.get("altitude"),
        values.get("rpm"),
        values.get("consumido"),
        values.get("pct_acelerado"),
        values.get("motor"),
        values.get("temperatura_motor"),
        values.get("ar_cond"),
        values.get("freio"),
        values.get("arla"),
        values.get("consumido_delta"),
        values.get("peso_total"),
    )


def sample_from_status(status: dict, state_name: str, is_new: bool, is_repeated: bool, source: str, chosen_reason: str):
    values = values_from_status(status)
    msg_tm = status.get("msg_tm")
    sample_tm = msg_tm if isinstance(msg_tm, (int, float)) else int(time.time())

    return {
        "sample_tm": int(sample_tm),
        "msg_tm": int(msg_tm) if isinstance(msg_tm, (int, float)) else None,
        "item_id": status.get("item_id"),
        "state": state_name,
        "is_gap": False,
        "is_new_sample": is_new,
        "is_repeated": is_repeated,
        "values": values,
        "source": source,
        "chosen_reason": chosen_reason,
    }


def build_status_from_sample(sample: dict, now_epoch: float):
    msg_tm = sample.get("msg_tm")
    age_sec = None
    if isinstance(msg_tm, (int, float)):
        age_sec = now_epoch - float(msg_tm)

    stale = False if age_sec is None else age_sec > STALE_THRESHOLD_SEC
    values = sample.get("values") or {}
    state_name = sample.get("state") or "new"

    return {
        "tm": msg_tm,
        "pos_tm": msg_tm,
        "msg_tm": msg_tm,
        "item_id": sample.get("item_id"),
        "age_sec": age_sec,
        "stale": stale,
        "unit_has_event": state_name != "stale",
        "snapshot_refreshed": bool(sample.get("is_new_sample")),
        "backend_snapshot_reloaded": sample.get("source") == "calc_last",
        "event_msg_tm": msg_tm,
        "has_new_sample": bool(sample.get("is_new_sample")),
        "has_fresh_data": not stale,
        "sample_state": state_name,
        "source": sample.get("source"),
        "chosen_reason": sample.get("chosen_reason"),
        "fields": {
            "velocidade": sensor_field(values.get("velocidade")),
            "altitude": sensor_field(values.get("altitude")),
            "rpm": sensor_field(values.get("rpm")),
            "consumido": sensor_field(values.get("consumido")),
            "pct_acelerado": sensor_field(values.get("pct_acelerado")),
            "motor": sensor_field(values.get("motor")),
            "temperatura_motor": sensor_field(values.get("temperatura_motor")),
            "ar_cond": sensor_field(values.get("ar_cond")),
            "freio": sensor_field(values.get("freio")),
            "arla": sensor_field(values.get("arla")),
            "consumido_delta": sensor_field(values.get("consumido_delta")),
            "peso_total": sensor_field(values.get("peso_total")),
        },
    }


def choose_best_status(avl_status: dict | None, calc_status: dict | None, item_id: int):
    last_msg_tm = get_last_msg_tm(item_id)
    last_signature = get_last_signature(item_id)

    avl_values = values_from_status(avl_status) if avl_status else None
    calc_values = values_from_status(calc_status) if calc_status else None

    avl_sig = build_values_signature(avl_values) if avl_values else None
    calc_sig = build_values_signature(calc_values) if calc_values else None

    avl_tm = avl_status.get("msg_tm") if avl_status else None
    calc_tm = calc_status.get("msg_tm") if calc_status else None

    avl_new = False
    calc_new = False

    if avl_status and avl_status.get("useful"):
        if avl_tm is not None and avl_tm != last_msg_tm:
            avl_new = True
        elif avl_sig is not None and avl_sig != last_signature:
            avl_new = True

    if calc_status:
        if calc_tm is not None and calc_tm != last_msg_tm:
            calc_new = True
        elif calc_sig is not None and calc_sig != last_signature:
            calc_new = True

    if avl_new:
        return avl_status, "avl_evts", "avl_changed"
    if calc_new:
        return calc_status, "calc_last", "calc_last_changed"
    if avl_status and avl_status.get("useful"):
        return avl_status, "avl_evts", "avl_same_snapshot"
    if calc_status:
        return calc_status, "calc_last", "calc_last_same_snapshot"
    return None, None, None


def maybe_append_repeated_sample(item_id: int, now_epoch: int):
    samples = list_live_samples(item_id, limit=1)
    if not samples:
        return None

    last_sample = samples[-1]
    if last_sample.get("is_gap"):
        return None

    last_msg_tm = last_sample.get("msg_tm")
    age_sec = None
    if isinstance(last_msg_tm, (int, float)):
        age_sec = now_epoch - float(last_msg_tm)

    if age_sec is None or age_sec > STALE_THRESHOLD_SEC:
        return None

    repeated = {
        "sample_tm": now_epoch,
        "msg_tm": last_msg_tm,
        "item_id": item_id,
        "state": "repeated",
        "is_gap": False,
        "is_new_sample": False,
        "is_repeated": True,
        "values": dict(last_sample.get("values") or {}),
        "source": "repeated",
        "chosen_reason": "no_change_after_avl_and_calc_last",
    }
    append_live_sample(item_id, repeated)
    set_last_plot_kind(item_id, "sample")
    return repeated


def maybe_append_gap(item_id: int, now_epoch: int):
    samples = list_live_samples(item_id, limit=1)
    last_msg_tm = None
    if samples:
        last_msg_tm = samples[-1].get("msg_tm")

    age_sec = None
    if isinstance(last_msg_tm, (int, float)):
        age_sec = now_epoch - float(last_msg_tm)

    if age_sec is None or age_sec <= STALE_THRESHOLD_SEC:
        return None

    if get_last_plot_kind(item_id) == "gap":
        return None

    gap = {
        "sample_tm": now_epoch,
        "msg_tm": last_msg_tm,
        "item_id": item_id,
        "state": "stale",
        "is_gap": True,
        "is_new_sample": False,
        "is_repeated": False,
        "values": {},
        "source": "gap",
        "chosen_reason": "stale_threshold_exceeded",
    }
    append_live_sample(item_id, gap)
    set_last_plot_kind(item_id, "gap")
    return gap


def bootstrap_from_calc_last(item_id: int):
    resp = wialon_ajax("unit/calc_last", build_params_obj())

    with STATE_LOCK:
        STATE["raw"] = resp

    if isinstance(resp, dict) and "error" in resp:
        raise RuntimeError(f"Wialon calc_last error code: {resp.get('error')}")

    status = build_status_from_calc_last(resp, source="calc_last", chosen_reason="bootstrap")
    sample = sample_from_status(
        status=status,
        state_name="bootstrap",
        is_new=True,
        is_repeated=False,
        source="calc_last",
        chosen_reason="bootstrap",
    )

    append_live_sample(item_id, sample)
    set_last_plot_kind(item_id, "sample")
    if sample["msg_tm"] is not None:
        set_last_msg_tm(item_id, sample["msg_tm"])
    set_last_signature(item_id, build_values_signature(sample["values"]))
    set_bootstrap_done(item_id, True)

    return build_status_from_sample(sample, time.time())


# -----------------------------
# Worker
# -----------------------------
def poll_loop():
    global AVL_CURSOR_TM

    while True:
        try:
            sid_now = get_sid()
            if not sid_now:
                raise RuntimeError("SID vazio. Atualize pelo dashboard.")

            item_id = get_item_id()

            with AVL_LOCK:
                tm_cursor = AVL_CURSOR_TM

            ev = wialon_avl_evts(tm_cursor)

            if isinstance(ev, dict) and isinstance(ev.get("tm"), (int, float)):
                with AVL_LOCK:
                    AVL_CURSOR_TM = int(ev["tm"])

            with STATE_LOCK:
                STATE["ev_raw"] = ev

            if isinstance(ev, dict) and "error" in ev:
                raise RuntimeError(f"Wialon avl_evts error code: {ev.get('error')}")

            now_epoch = time.time()

            if not get_bootstrap_done(item_id):
                existing = list_live_samples(item_id, limit=1)
                if not existing:
                    status = bootstrap_from_calc_last(item_id)
                    with STATE_LOCK:
                        STATE["ok"] = True
                        STATE["last_update_epoch"] = now_epoch
                        STATE["error"] = None
                        STATE["data"] = status
                    time.sleep(POLL_SECONDS)
                    continue
                set_bootstrap_done(item_id, True)

            avl_status = build_status_from_avl(ev if isinstance(ev, dict) else {}, item_id)

            calc_resp = wialon_ajax("unit/calc_last", build_params_obj())
            with STATE_LOCK:
                STATE["raw"] = calc_resp

            if isinstance(calc_resp, dict) and "error" in calc_resp:
                raise RuntimeError(f"Wialon calc_last error code: {calc_resp.get('error')}")

            calc_status = build_status_from_calc_last(
                calc_resp,
                source="calc_last",
                chosen_reason="cycle_reconciliation",
            )

            chosen_status, source, chosen_reason = choose_best_status(
                avl_status=avl_status,
                calc_status=calc_status,
                item_id=item_id,
            )

            latest_status = None

            if chosen_status is not None:
                is_new = chosen_reason in {"avl_changed", "calc_last_changed"}
                if is_new:
                    sample = sample_from_status(
                        status=chosen_status,
                        state_name="new",
                        is_new=True,
                        is_repeated=False,
                        source=source,
                        chosen_reason=chosen_reason,
                    )
                    append_live_sample(item_id, sample)
                    set_last_plot_kind(item_id, "sample")
                    if sample["msg_tm"] is not None:
                        set_last_msg_tm(item_id, sample["msg_tm"])
                    set_last_signature(item_id, build_values_signature(sample["values"]))
                    latest_status = build_status_from_sample(sample, now_epoch)
                else:
                    repeated = maybe_append_repeated_sample(item_id, int(now_epoch))
                    if repeated is not None:
                        latest_status = build_status_from_sample(repeated, now_epoch)
                    else:
                        gap = maybe_append_gap(item_id, int(now_epoch))
                        if gap is not None:
                            latest_status = build_status_from_sample(gap, now_epoch)
                        else:
                            samples = list_live_samples(item_id, limit=1)
                            if samples:
                                latest_status = build_status_from_sample(samples[-1], now_epoch)
            else:
                repeated = maybe_append_repeated_sample(item_id, int(now_epoch))
                if repeated is not None:
                    latest_status = build_status_from_sample(repeated, now_epoch)
                else:
                    gap = maybe_append_gap(item_id, int(now_epoch))
                    if gap is not None:
                        latest_status = build_status_from_sample(gap, now_epoch)
                    else:
                        samples = list_live_samples(item_id, limit=1)
                        if samples:
                            latest_status = build_status_from_sample(samples[-1], now_epoch)

            if latest_status is None:
                latest_status = {
                    "tm": None,
                    "pos_tm": None,
                    "msg_tm": None,
                    "item_id": item_id,
                    "age_sec": None,
                    "stale": True,
                    "unit_has_event": False,
                    "snapshot_refreshed": False,
                    "backend_snapshot_reloaded": False,
                    "event_msg_tm": None,
                    "has_new_sample": False,
                    "has_fresh_data": False,
                    "sample_state": "stale",
                    "source": None,
                    "chosen_reason": "no_data_available",
                    "fields": {},
                }

            with STATE_LOCK:
                STATE["ok"] = True
                STATE["last_update_epoch"] = now_epoch
                STATE["error"] = None
                STATE["data"] = latest_status

        except Exception as e:
            with STATE_LOCK:
                STATE["ok"] = False
                STATE["last_update_epoch"] = time.time()
                STATE["error"] = str(e)

        time.sleep(POLL_SECONDS)


# -----------------------------
# HTTP Server
# -----------------------------
class Handler(BaseHTTPRequestHandler):
    def _set_common_headers(self, content_type: str, content_length: int | None = None):
        self.send_header("Content-Type", content_type)
        if content_length is not None:
            self.send_header("Content-Length", str(content_length))

        allowed_origin = get_allowed_origin(self)
        self.send_header("Access-Control-Allow-Origin", allowed_origin)
        self.send_header("Access-Control-Allow-Methods", CORS_ALLOW_METHODS)
        self.send_header("Access-Control-Allow-Headers", CORS_ALLOW_HEADERS)
        self.send_header("Access-Control-Max-Age", "86400")
        self.send_header("Vary", "Origin")

    def _send_bytes(self, content: bytes, content_type: str, code=200):
        self.send_response(code)
        self._set_common_headers(content_type, len(content))
        self.end_headers()
        self.wfile.write(content)

    def _send_text(self, text: str, content_type: str, code=200):
        self._send_bytes(text.encode("utf-8"), content_type, code)

    def _send_json(self, obj, code=200):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self._set_common_headers("application/json; charset=utf-8", len(body))
        self.end_headers()
        self.wfile.write(body)

    def _send_html(self, html: str, code=200):
        self._send_text(html, "text/html; charset=utf-8", code)

    def _read_raw_body(self):
        length = int(self.headers.get("Content-Length", "0"))
        return self.rfile.read(length) if length > 0 else b""

    def _safe_json_header(self, name, default=None):
        raw = self.headers.get(name)
        if not raw:
            return default
        try:
            return json.loads(raw)
        except Exception:
            return default

    def _serve_file(self, file_path: Path, content_type: str | None = None, not_found_msg: str = "arquivo nao encontrado"):
        try:
            if not file_path.exists() or not file_path.is_file():
                self._send_text(not_found_msg, "text/plain; charset=utf-8", 404)
                return
            self._send_bytes(file_path.read_bytes(), content_type or guess_content_type(file_path), 200)
        except Exception as e:
            self._send_text(f"Erro ao servir arquivo: {e}", "text/plain; charset=utf-8", 500)

    def _serve_static_from_frontend(self, request_path: str):
        file_path = resolve_static_path(request_path)
        if file_path is None:
            self._send_json({"error": "not_found"}, 404)
            return

        if not file_path.exists() or not file_path.is_file():
            self._send_json({"error": "not_found"}, 404)
            return

        self._serve_file(file_path, guess_content_type(file_path), "arquivo nao encontrado")

    def do_OPTIONS(self):
        allowed_origin = get_allowed_origin(self)
        if allowed_origin == "null":
            self.send_response(403)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.send_header("Content-Length", "0")
            self.send_header("Vary", "Origin")
            self.end_headers()
            return

        self.send_response(204)
        self._set_common_headers("text/plain; charset=utf-8", 0)
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        qs = parse_qs(parsed.query)

        if path == "/health":
            self._send_json({"ok": True, "service": "dashboard_collector"}, 200)
            return

        if path == "/status":
            with STATE_LOCK:
                payload = {
                    "ok": STATE["ok"],
                    "last_update_epoch": STATE["last_update_epoch"],
                    "error": STATE["error"],
                    "data": STATE["data"],
                }
            self._send_json(payload, 200)
            return

        if path == "/live_series":
            try:
                item_id = get_item_id()
                limit_raw = (qs.get("limit") or ["180"])[0]
                try:
                    limit = max(1, min(int(limit_raw), LIVE_BUFFER_SIZE))
                except Exception:
                    limit = 180

                samples = list_live_samples(item_id, limit=limit)
                self._send_json(
                    {
                        "ok": True,
                        "item_id": item_id,
                        "count": len(samples),
                        "samples": samples,
                    },
                    200,
                )
            except Exception as e:
                self._send_json({"ok": False, "error": str(e), "samples": []}, 500)
            return

        if path == "/raw":
            with STATE_LOCK:
                payload = STATE.get("raw")
            self._send_json({"raw": payload}, 200)
            return

        if path == "/ev_raw":
            with STATE_LOCK:
                payload = STATE.get("ev_raw")
            self._send_json({"ev_raw": payload}, 200)
            return

        if path == "/units":
            try:
                sid_now = get_sid()
                if not sid_now:
                    self._send_json({"error": "SID vazio"}, 400)
                    return

                params = {
                    "spec": {
                        "itemsType": "avl_unit",
                        "propName": "sys_name",
                        "propValueMask": "*",
                        "sortType": "sys_name",
                    },
                    "force": 1,
                    "flags": 1,
                    "from": 0,
                    "to": 0,
                }

                resp = wialon_ajax("core/search_items", params)

                if isinstance(resp, dict) and "error" in resp:
                    code = resp.get("error")
                    if code == 1:
                        self._send_json({"error": "SID inválido/expirado. Atualize o SID."}, 401)
                        return
                    self._send_json({"error": f"Wialon error code: {code}"}, 400)
                    return

                items = (resp or {}).get("items", [])
                units = [
                    {"id": it.get("id"), "name": it.get("nm")}
                    for it in items
                    if it.get("id") and it.get("nm")
                ]
                self._send_json({"units": units}, 200)

            except Exception as e:
                self._send_json({"error": str(e)}, 500)
            return

        if path == "/treatment_status":
            job_id = (qs.get("job_id") or [""])[0].strip()
            if not job_id:
                self._send_json({"ok": False, "error": "missing job_id"}, 400)
                return

            status = get_treatment_status(job_id)
            if not status:
                self._send_json({"ok": False, "error": "job_not_found"}, 404)
                return

            self._send_json(
                {
                    "ok": True,
                    "job_id": job_id,
                    "status": status.get("status"),
                    "progress": status.get("progress"),
                    "result": status.get("result"),
                    "error": status.get("error"),
                },
                200,
            )
            return

        if path == "/download_treatment_result":
            job_id = (qs.get("job_id") or [""])[0].strip()
            if not job_id:
                self._send_json({"ok": False, "error": "missing job_id"}, 400)
                return

            status = get_treatment_status(job_id)
            if status and status.get("status") != "done":
                self._send_json({"ok": False, "error": "job_not_ready"}, 409)
                return

            job = get_treatment_job(job_id)
            if not job:
                self._send_json({"ok": False, "error": "job_not_found"}, 404)
                return

            file_path = Path(job["path"])
            if not file_path.exists():
                self._send_json({"ok": False, "error": "file_not_found"}, 404)
                return

            content = file_path.read_bytes()
            self.send_response(200)
            self._set_common_headers(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                len(content),
            )
            self.send_header(
                "Content-Disposition",
                f'attachment; filename="{job["download_name"]}"'
            )
            self.end_headers()
            self.wfile.write(content)
            return

        if path in ("/", "/dashboard.html"):
            self._send_html(load_dashboard_html(), 200)
            return

        self._serve_static_from_frontend(path)

    def do_POST(self):
        path = urlparse(self.path).path

        if path in {"/process_treatment", "/process_treatment_step1"}:
            try:
                raw_body = self._read_raw_body()
                if not raw_body:
                    self._send_json({"ok": False, "error": "empty body"}, 400)
                    return

                raw_filename = self.headers.get("X-Filename", "arquivo.xlsx")
                try:
                    filename = unquote(raw_filename)
                except Exception:
                    filename = raw_filename

                job_id = uuid.uuid4().hex
                input_path = save_treatment_input_file(raw_body, filename, job_id)

                if path == "/process_treatment":
                    create_treatment_status(job_id, "base", filename, input_path)
                    worker = threading.Thread(
                        target=run_treatment_job,
                        args=(job_id, input_path, filename),
                        daemon=True,
                    )
                else:
                    config = self._safe_json_header("X-Treatment-Config", {}) or {}
                    config = {
                        "bestRpmMin": float(config.get("bestRpmMin", 1100)),
                        "bestRpmMax": float(config.get("bestRpmMax", 1900)),
                        "speedLowMax": float(config.get("speedLowMax", 10)),
                        "speedMediumMax": float(config.get("speedMediumMax", 15)),
                        "rpmUsefulStart": float(config.get("rpmUsefulStart", 900)),
                        "rpmLightStart": float(config.get("rpmLightStart", 1100)),
                        "rpmUsefulEnd": float(config.get("rpmUsefulEnd", 1900)),
                        "accelLightMax": float(config.get("accelLightMax", 30)),
                        "accelMediumMax": float(config.get("accelMediumMax", 60)),
                        "brakeMediumMin": float(config.get("brakeMediumMin", 2)),
                        "brakeIntenseMin": float(config.get("brakeIntenseMin", 4)),
                    }
                    create_treatment_status(job_id, "step1", filename, input_path)
                    worker = threading.Thread(
                        target=run_treatment_step1_job,
                        args=(job_id, input_path, filename, config),
                        daemon=True,
                    )

                worker.start()
                self._send_json({"ok": True, "job_id": job_id, "status": "queued"}, 202)
                return

            except Exception as e:
                self._send_json({"ok": False, "error": str(e)}, 500)
                return

        try:
            length = int(self.headers.get("Content-Length", "0"))
            raw_body = self.rfile.read(length) if length > 0 else b"{}"
            body = json.loads(raw_body.decode("utf-8"))
        except Exception:
            self._send_json({"ok": False, "error": "invalid json"}, 400)
            return

        if path == "/set_sid":
            new_sid = (body.get("sid") or "").strip()
            if not new_sid:
                self._send_json({"ok": False, "error": "missing sid"}, 400)
                return

            set_sid(new_sid)
            reset_runtime_state(get_item_id())

            self._send_json({"ok": True}, 200)
            return

        if path == "/set_unit":
            item_id = body.get("itemId")
            try:
                item_id = int(item_id)
            except Exception:
                self._send_json({"ok": False, "error": "invalid itemId"}, 400)
                return

            if item_id <= 0:
                self._send_json({"ok": False, "error": "invalid itemId"}, 400)
                return

            set_item_id(item_id)
            reset_runtime_state(item_id)

            self._send_json({"ok": True}, 200)
            return

        if path == "/login_fonte":
            self._send_json(
                {
                    "ok": False,
                    "error": "login_fonte ainda nao implementado. Preciso do request real de login da plataforma fonte para montar essa rota corretamente."
                },
                501,
            )
            return

        self._send_json({"error": "not_found"}, 404)

    def log_message(self, format, *args):
        return


def main():
    print(f"[PY] Executando: {Path(__file__).resolve()}", flush=True)
    print(f"[PY] Frontend dir: {FRONTEND_DIR}", flush=True)
    print(f"[PY] HTML esperado: {HTML_PATH}", flush=True)
    print(f"[PY] CSS esperado: {CSS_PATH}", flush=True)
    print(f"[PY] Logo esperada: {LOGO_PATH}", flush=True)
    print(f"[PY] Host: {HTTP_HOST}:{HTTP_PORT}", flush=True)
    print(f"[PY] CORS allowed origins: {CORS_ALLOWED_ORIGINS}", flush=True)

    t = threading.Thread(target=poll_loop, daemon=True)
    t.start()

    server = ThreadingHTTPServer((HTTP_HOST, HTTP_PORT), Handler)
    print(f"OK: servidor em http://localhost:{HTTP_PORT}/", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
